from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import datetime
import pandas as pd
from openpyxl import Workbook
import logging

PATH = "chromedriver.exe"
service = ChromeService(executable_path=PATH)
driver = webdriver.Chrome(service=service)
logging.basicConfig(filename='error_messages.log',
                    encoding='utf-8', level=logging.WARNING, format='%(asctime)s:%(levelname)s:%(message)s')

# get header and data id from html site of the lva
# see README.md for more explanation
id = 'j_id_2p:j_id_92'

def read_lva(lva_number, semester_jahr=datetime.date.today().year, semester_symbol='w'):
    """ Read exam dates from TISS

    Parameters: lva_number (str) -> TISS LVA Number  e.g. 303.009 or 303009 
                semester_jahr (str) -> Semester where the exams should be optained (optional)
                semester_symbol (str) -> Either winter (W) or summer (S)

    Returns: df (datafrane) -> each row represents a single exam
    """
    # create exam list
    exam_lst = []
    # create header list
    header = []
    # clean LVA number input
    lva_number = lva_number.replace(".", "")
    # base url for the lva
    base_url = f"https://tiss.tuwien.ac.at/course/courseDetails.xhtml?courseNr={lva_number}&semester={semester_jahr}{semester_symbol}"
    # get url content
    driver.get(base_url)
    # puffer for slow browsers max waiting 5 sec
    wait = WebDriverWait(driver, 5)
    try:
        __ = wait.until(EC.presence_of_element_located(
            [By.XPATH, f"//*[@id='{str(id)}_data']"]))
    except TimeoutException as exception:
        pass
    # get table header length
    header_length = len(driver.find_elements(
        "xpath", f"//*[@id='{str(id)}_head']/tr[1]/th"))
    print('length')
    print(header_length)
    # loop over header
    for title in range(1, header_length+1):
        value = driver.find_element(
            "xpath", f"//*[@id='{str(id)}_head']/tr[1]/th[{str(title)}]").text
        header.append(value)
    # count number of rows
    rows = len(driver.find_elements(
        "xpath", f"//*[@id='{str(id)}_data']/tr"))
    # count number of columns
    columns = len(driver.find_elements(
        "xpath", f"//*[@id='{str(id)}_data']/tr[1]/td"))
    # loop over the table
    for r in range(1, rows+1):
        single_exam = []
        for c in range(1, columns+1):
            value = driver.find_element(
                "xpath", f"//*[@id='{str(id)}_data']/tr[{str(r)}]/td[{str(c)}]").text
            single_exam.append(value)
        exam_lst.append(single_exam)
    # get LVA name
    name = driver.title
    # strip title of non name charaters
    for r in ((".", ""), ("| TU Wien", ""), (str(lva_number), "")):
        name = name.replace(*r)
    # strip leading whitespaces
    name = name.strip()
    # create pandas dataframe
    df = pd.DataFrame(exam_lst, columns=header)
    # name dataframe
    df.name = name
    if not exam_lst or not header:
        logging.warning(
            f"Either no exams are specified or LVA-Number is wrong! (LVA-NR = {lva_number}, URL = {base_url})")
    return df


def get_lva_numbers(file, semester):
    dic = pd.read_excel(file, sheet_name=[semester], usecols="B")
    df = dic[semester]
    lst = df.loc[:, "LVA Nummer"].to_list()
    return lst


def create_new_file(source_file):
    xl = pd.ExcelFile(source_file)
    sheet_names = xl.sheet_names
    workbook = Workbook()
    for sheet in sheet_names:
        workbook.create_sheet(title=sheet)
    del workbook["Sheet"]
    workbook.save("Termine.xlsx")


def main():
    # chrome setup for selenium
    excel_file = "LVA-Nummern.xlsx"
    # create empty xlsx file
    create_new_file(excel_file)
    xl = pd.ExcelFile(excel_file)
    sheet_names = xl.sheet_names
    # loop over sheetnames for semester name
    for sheet in sheet_names:
        lva_numbers = get_lva_numbers(excel_file, sheet)
    # loop over lvas
        for number in lva_numbers:
            number = str(number)
            df = read_lva(number, semester_symbol='W')
            if df.empty:
                continue
            data = df.loc[:, ["Zeit", "Datum", "Ort"]]
            # rename index
            data.index.name = df.name
            # shift index to start at 1
            data.index += 1
            # write data to excel file
            with pd.ExcelWriter("Termine.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                data.to_excel(writer, sheet_name=sheet,
                              startrow=writer.sheets[sheet].max_row+1)
    # close the driver
    driver.quit()


if __name__ == "__main__":
    main()
